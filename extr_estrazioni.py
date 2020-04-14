#!/usr/bin/env python3

import mysql.connector
import openpyxl
import argparse
import glob
import sys

parser = argparse.ArgumentParser()
parser.add_argument('--extr_folder', help='foo help')
args = parser.parse_args()


mydb = mysql.connector.connect(
  host="colab1.cibio.unitn.it",
  user="covid_user",
  passwd="Q2GtXNpnKj94IP4HEo0IyvCun",
  database="covid",
  port=33006
)


samplesToAdd=[]

for extratcionsFile in glob.glob(args.extr_folder+'/*.xls*'):
	wb_obj = openpyxl.load_workbook(extratcionsFile,data_only=True,read_only=True) 
	sheet = wb_obj.active

	print(extratcionsFile,sheet.max_row, "rows", sheet.max_column, "cols")


	batchName= sheet["B2"].value.split('_')[0]
	batchDate= sheet["B2"].value.split('_')[1]

	realDate = '20'+batchDate[0:2]+'-'+batchDate[2:4]+'-'+batchDate[4:6]


	for e in range(7,30):
		if sheet["B"+str(e)].value is not None:
			barcode= str(sheet["B"+str(e)].value)[0:-2]

		 
			samplesToAdd.append( (barcode, realDate,batchName) )
			print (extratcionsFile,barcode, realDate,batchName)


mycursor = mydb.cursor()
sql = "INSERT IGNORE INTO estrazioni (barcode, data_estrazione, batch) VALUES (%s, %s,%s)"
mycursor.executemany(sql, samplesToAdd)

sql = "INSERT IGNORE INTO samples (barcode, data_checkin) VALUES (%s, %s)"
mycursor.executemany(sql, [(_[0],_[1]) for _ in samplesToAdd ] )




mydb.commit()