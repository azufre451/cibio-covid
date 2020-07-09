#!/usr/bin/env python3

import mysql.connector
import openpyxl
import argparse
import glob
import os
import sys
from collections import defaultdict
from locale import atof, setlocale, LC_NUMERIC, LC_ALL
from itertools import islice
import pandas as pd
from conf import DBConf
#setlocale(LC_NUMERIC,('it_IT','UTF-8'))



def na2none(a):

	if a == 'N/A':
		return None
	else:

		return float(a)


parser = argparse.ArgumentParser()
parser.add_argument('--data_folder', help='his is the folder containing all the Analisi Excell files. One file per plate.', required=True)
parser.add_argument('--control_wells', help="manual setup for control wells in the plate (list, space separated. Esample: A01 A02 A03 ...)", default=[], nargs='+')
#parser.add_argument('--kit',default="bosphore")
parser.add_argument('--replace_data', action='store_true',  help="overwrite the data")
parser.add_argument('--dry_run', action='store_true',  help="Run without actually updating the database.")
#parser.add_argument('--kf', action='store_true',help='Equals to --control_wells A01 A02 A12 D04 D08')
#parser.add_argument('--old_plates', action='store_true',help='Equals to --control_wells H01 A01 A02 A05 A08 A12')
parser.add_argument('--well_allow', default=[], nargs='+', help="Allows to override the plate setup to include a sample in an otherwise 'control' well")
parser.add_argument('--platename_from_file',action='store_true', help='allows to take the Plate Name from the filename, instead thatn from the designated cell in the template')




args = parser.parse_args()

#control_samples = ['NTC','PK','PE','BE','55555','77777','11111111','33333','20202020','25252525','27272727', 'pcr NEG','pcr POS']

fluorophores=["FAM","HEX", "Texas Red","Cy5"]

#if args.kf:
#	ListOfControlWells=['A01','A02','A12','D04','D08']
#elif args.old_plates:
#	ListOfControlWells=['H01','A01','A02','A05','A08','A12']
#else:
#	ListOfControlWells=list(args.control_wells)

#WellsToAvoid = [_ for _ in ListOfControlWells if _ not in list(args.well_allow)]

#print("Skipping wells: ",' '.join(WellsToAvoid))
#print("-- Pozzetti di Controllo: "+' '.join(WellsToAvoid))

try:


	mydb = mysql.connector.connect(
		host=DBConf.colab_host,
		user=DBConf.colab_user,
		passwd=DBConf.colab_passwd,
		database=DBConf.colab_database,
		port=DBConf.colab_port,
		auth_plugin=DBConf.colab_auth_plugin
	)

	mycursor = mydb.cursor()
	print("-- Connessione al Database riuscita")

except mysql.connector.Error as err:
	print("-- <span class=\"errorMessage\">ERRORE</span> del Database: ", str(err))


barcode2pool={}
pool2barcode=defaultdict(set)
estrazioniToAdd=[]
samplesToAdd=[]
curves2add=[]
samplesToCheck=[]
results_tracker=defaultdict(int)

for analFile in glob.glob(args.data_folder+'/*.xls*'):

	wb_obj = openpyxl.load_workbook(analFile,data_only=True, read_only=True)
	wells_with_data = set()

	try:
		sheet = wb_obj["analysis"]
	except:
		#print("Error in finding Data tab", analFile )
		print("-- ERRORE: il file non ha la tab 'analysis'")
		sys.exit(1)

	has_estrazioni = "report_analysis" in wb_obj

	if has_estrazioni:
		report_sheet = wb_obj["report_analysis"]
		hh={}
		for row in report_sheet.iter_rows():
			if len(hh) == 0:
				hh = {field.value: pos for (pos, field) in enumerate(row)}
				continue
			if not int(row[hh['is_empty']].value):
				barcode               = str(row[hh['barcode']].value)
				pool_barcode          = str(row[hh['pool_barcode']].value)
				barcode2pool[barcode] = pool_barcode
				pool2barcode[pool_barcode].add(barcode)


	hh={}
	plateName=None
	for row in sheet.iter_rows():
		if len(hh) == 0:
			hh = {field.value: pos for (pos, field) in enumerate(row)}

			continue

		if plateName is None:
			if not args.platename_from_file:
				plateName= row[hh['plate']].value

				print("-- Apertura: <span class=\"emph\">", os.path.basename(analFile),'</span>')

				if plateName == 'XXXy' and analFile == 'Analisi/200408/V220040801_analisi.xlsx': plateName = 'V220040801'
			else:
				plateName = os.path.basename(analFile).replace('_analisi.xlsm','')

			plateDate= '20'+plateName[2:4]+'-'+plateName[4:6]+'-'+plateName[6:8]


		well_nozero= str(row[hh['well_no_zero']].value)
		well= str(row[hh['well_with_zero']].value)

		barcode= str(row[hh['barcode']].value)
		batch_kf=str(row[hh['batch']].value)

		if has_estrazioni:
			batch_id  = str(row[hh['batch_kf']].value)
			batchName = batch_id.split('_')[0]
			batchDate = batch_id.split('_')[1]
			realDate  = '20'+batchDate[0:2]+'-'+batchDate[2:4]+'-'+batchDate[4:6]

		val_cy5 = na2none(row[hh['cy5_cq']].value)
		val_fam = na2none(row[hh['fam_cq']].value)
		val_hex = na2none(row[hh['hex_cq']].value)
		val_tred = na2none(row[hh['texas_red_cq']].value)
		isempty = int(row[hh['is_well_empy']].value)
		is_control = int(row[hh['is_control']].value) # int(barcode in control_samples or well in WellsToAvoid)
		is_pool = int(row[hh['is_pool']].value) if 'is_pool' in hh else False

		auto_result = row[hh['test_result_auto']].value

		final_result = row[hh['final_result']].value if row[hh['final_result']].value != 'NON REFERTARE' else row[hh['user_result']].value

		kit=str(row[hh['kit_name']].value)

		if final_result not in ['POSITIVO','NEGATIVO','RIPETERE ESTRAZIONE','RIPETERE PCR','RIPETERE TAMPONE']:
			if is_control:
				final_result = 'CONTROLLO'
				auto_result = 'CONTROLLO'
			elif final_result == 'CORRETTO' and auto_result == 'RIPETERE TAMPONE':
				final_result = 'RIPETERE TAMPONE'
			else:
				final_result = 'ERRORE COMPILAZIONE'

		if str (barcode) != "0" and barcode != 'None' and  not isempty:
			result = False

			if has_estrazioni:
				if is_pool:
					result = barcode in pool2barcode
				else:
					result = barcode in barcode2pool
			else:
				sql = 'SELECT 1 FROM samples WHERE barcode = %s'
				mycursor.execute(sql, (barcode,))
				mycursor.fetchone()
				result = mycursor.rowcount > 0

			generate_row_data = lambda real_barcode, real_pool_barcode: (plateName, plateDate,real_barcode,real_pool_barcode,well,val_cy5,val_fam,val_hex,val_tred,auto_result,final_result, is_control,batch_kf,kit)

			barcodes_in_well = pool2barcode[barcode] if is_pool else {barcode}

			for bb in barcodes_in_well:
				row_data = generate_row_data(bb, barcode)
				# print("OK", *row_data)
				if result:
					if has_estrazioni: estrazioniToAdd.append( (bb, realDate, batchName) )
					samplesToAdd.append(row_data)
					results_tracker[final_result] += 1
					# print("OK", *estrazioniToAdd[len(estrazioniToAdd) - 1])
				elif str(barcode) != "0":
					samplesToCheck.append(row_data)

			if result:
				wells_with_data.add(well_nozero)

			#if final_result not in ['POSITIVO','NEGATIVO','RIPETERE ESTRAZIONE','RIPETERE PCR','RIPETERE TAMPONE']:
				#print("NOK")
	if(len(samplesToAdd) > 0):
		print("-- ",len(samplesToAdd)," Campioni da aggiungere")
		for result,counter in results_tracker.items():
			print("----- ",counter," Campioni con esito ",result.lower())


	if(len(samplesToCheck) > 0):
		print("-- ",len(samplesToCheck)," Campioni con errori | <span class=\"errorMessage\">ATTENZIONE</span>")
		for k in samplesToCheck:
			print("----- Barcode: ",k[2], ', pozzetto ',k[3],' (',k[8],')')



	for fluorophore in fluorophores:

		try:
			sheetFam = wb_obj["Curve "+fluorophore]
		except:
			print("-- <span class=\"errorMessage\">ERRORE</span> - Non trovo la curva di ",fluorophore)


			sys.exit(0)

		data = sheetFam.values
		cols = next(data)[1:]
		data = list(data)
		idx = [r[0] for r in data]
		data = (islice(r, 1, None) for r in data)
		df = pd.DataFrame(data, index=idx, columns=cols).set_index('Cycle')


		for wellTo in df.columns:
			if wellTo in wells_with_data:
				wellDB = wellTo[0]+wellTo[1:].zfill(2)
				curveString = ','.join(map(str,list(df[wellTo])))
				curves2add.append( (plateName,wellDB,fluorophore,curveString) )
				# print("OK", *curves2add[len(curves2add) - 1])

if not args.dry_run:
	try:

		if args.replace_data:
			sql = 'DELETE FROM pcr_plates WHERE plate  = %s'
			mycursor.execute(sql, (plateName,))
			sql = 'DELETE FROM curves WHERE plate  = %s'
			mycursor.execute(sql, (plateName,))
			print("-- Eliminazione dati vecchia plate | <span class=\"okMessage\">OK</span>")

		if len(estrazioniToAdd) > 0:
			sql = "INSERT IGNORE INTO samples (barcode, data_checkin) VALUES (%s, %s)"
			mycursor.executemany(sql, [(_[0],_[1]) for _ in estrazioniToAdd ] )

			print("-- "+str(mycursor.rowcount)+" Nuovi Campioni caricati")

			sql = "INSERT IGNORE INTO estrazioni (barcode, data_estrazione, batch) VALUES (%s, %s,%s)"
			mycursor.executemany(sql, estrazioniToAdd)

			print("-- "+str(mycursor.rowcount)+" Nuove Estrazioni caricate")

			print("-- Caricamento Campioni | <span class=\"okMessage\">OK</span>")

		sql = 'INSERT IGNORE INTO pcr_plates (plate, data_pcr, barcode, pooled_barcode, well, Cy5, FAM, HEX, TRed, esito_automatico, esito_pcr, isControl,batch_kf,kit) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
		mycursor.executemany(sql, samplesToAdd)
		print("-- Caricamento Curve | <span class=\"okMessage\">OK</span>")


		sql = 'INSERT IGNORE INTO curves (plate, well, fluorophore, curve) VALUES (%s,%s,%s,%s)'
		mycursor.executemany(sql, curves2add)
		print("-- Caricamento Plate | <span class=\"okMessage\">OK</span>")


		mydb.commit()
		if len(samplesToCheck) == 0:
			print("-- File caricato con successo, nessun errore! | <span class=\"okMessage\">OK</span>")
		else:
			print("-- File caricato con errori! | <span class=\"errorMessage\">ATTENZIONE</span>")

	except mysql.connector.Error as err:
		print("-- <span class=\"errorMessage\">ERRORE</span> del Database: ", str(err))
else:
	print("Dry run enabled: Not updating the database.")
